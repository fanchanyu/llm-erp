#!/usr/bin/env python3
"""
LLM-ERP 客戶 Email 批量查詢工具
用法：python3 lookup_emails.py
從 Excel 讀取「need research」的公司，批次搜尋網站並更新 email。
"""
import openpyxl, os, json, re, time, urllib.parse, urllib.request
from openpyxl.styles import PatternFill

SALES = os.path.dirname(os.path.abspath(__file__))
INPUT = os.path.join(SALES, "LLM-ERP客戶潛在名單.xlsx")
CACHE = os.path.join(SALES, "domain_cache.json")

# ─── 已知手動補完的 domain ──────────────────────────────────────
MANUAL_DOMAINS = {
    # 可自行擴充： "公司關鍵字": "domain.com.tw",
}

# ─── Google 自訂搜尋（如需使用請填入 API Key） ──────────────────
# 免費版每天 100 次查詢
GOOGLE_API_KEY = ""     # 填入你的 API Key
GOOGLE_CX = ""          # 填入你的 Search Engine ID

# ─── 從公司名推斷 domain ────────────────────────────────────────
def guess_domain_from_name(name):
    """Generate likely domain from company name patterns."""
    # Remove common suffixes
    n = name.replace(" ", "").replace(" ", "").replace("股份有限公司", "").replace("有限公司", "")
    n = n.replace("實業", "").replace("企業", "").replace("工業", "").replace("科技", "")
    n = n.replace("精密", "").replace("機械", "").replace("電子", "").replace("半導體", "")
    
    # Try common Taiwan patterns
    patterns = [
        lambda x: re.sub(r'[^a-zA-Z]', '', x).lower() + ".com.tw",
        lambda x: "www." + re.sub(r'[^a-zA-Z]', '', x).lower() + ".com.tw",
    ]
    
    # Convert Chinese to pinyin initials? Too complex.
    # Return None — let Google search handle it.
    return None

# ─── 使用 Google 搜尋查 domain ──────────────────────────────────
def search_domain_google(company_name, api_key, cx):
    """Search Google for company website."""
    if not api_key or not cx:
        return None
    query = urllib.parse.quote(f"{company_name} 官網 OR 官方網站")
    url = f"https://www.googleapis.com/customsearch/v1?key={api_key}&cx={cx}&q={query}&hl=zh-TW"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read().decode())
        for item in data.get("items", []):
            link = item.get("link", "")
            # Filter to likely company domains
            if any(ext in link for ext in [".com.tw", ".com", ".tw"]):
                from urllib.parse import urlparse
                domain = urlparse(link).netloc
                if domain.startswith("www."):
                    domain = domain[4:]
                return domain
        return None
    except Exception as e:
        return None

# ─── 主流程 ──────────────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(INPUT)
    
    # Load cache
    cache = {}
    if os.path.exists(CACHE):
        with open(CACHE, "r") as f:
            cache = json.load(f)
    
    total_updated = 0
    
    for ws in wb.worksheets:
        if ws.title == "總計":
            continue
        
        print(f"\n🔍 {ws.title}: 檢查未驗證的公司...")
        
        for row in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row, column=13)  # Email狀態
            if status_cell.value and str(status_cell.value).strip() == "verified":
                continue  # 已驗證跳過
            
            name = str(ws.cell(row=row, column=1).value or "").strip()
            if not name:
                continue
            
            # Check manual domains
            domain = None
            for kw, d in MANUAL_DOMAINS.items():
                if kw in name:
                    domain = d
                    break
            
            # Check cache
            if not domain and name in cache:
                domain = cache[name]
            
            # Google search
            if not domain:
                domain = search_domain_google(name, GOOGLE_API_KEY, GOOGLE_CX)
                if domain:
                    cache[name] = domain
                    time.sleep(0.5)  # Rate limit
            
            if domain:
                ws.cell(row=row, column=11).value = f"https://www.{domain}"
                ws.cell(row=row, column=12).value = f"info@{domain}"
                status_cell.value = "verified"
                total_updated += 1
                print(f"  ✅ {name[:20]:20s} → {domain}")
    
    # Save cache
    with open(CACHE, "w") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)
    
    # Save Excel
    wb.save(INPUT)
    
    # Summary
    verified = 0
    need_research = 0
    for ws in wb.worksheets:
        if ws.title == "總計":
            continue
        for row in range(2, ws.max_row + 1):
            status = str(ws.cell(row=row, column=13).value or "")
            if "verified" in status:
                verified += 1
            else:
                need_research += 1
    
    print(f"\n{'='*50}")
    print(f"✅ 本次新增：{total_updated} 家")
    print(f"📊 總計：{verified} verified + {need_research} need research = {verified+need_research}")
    print(f"\n💡 使用方法：")
    print(f"  1. 在 Google Cloud Console 啟用 Custom Search API")
    print(f"  2. 填入上方 GOOGLE_API_KEY 和 GOOGLE_CX")
    print(f"  3. 執行 python3 lookup_emails.py")
    print(f"  4. 或直接打電話查詢 — 電話號碼已在 Excel 中")
    print(f"\n📁 快取已儲存: {CACHE}")

if __name__ == "__main__":
    main()
